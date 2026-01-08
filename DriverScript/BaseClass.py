import configparser
import sys
import cx_Oracle
import pandas as pd
from datetime import datetime
import logging
import os
import openpyxl
from pathlib import Path


class BaseClass:
    project_path=str(Path.cwd().parent)
    print(project_path)
    def __init__(self):
        self.conn=None
        #
        os.environ['project_path'] = self.project_path
        self.initialize_variable()
        self.testCaseList=self.get_testcase_list()

    def initialize_variable(self):
        try:
            execution_path =os.environ['project_path']+'\\Configurations\\TestExecution.xlsx'
            wb = openpyxl.load_workbook(execution_path)
            ws = wb['GlobalConfig']
            os.environ['BrowserType'] = ws['A2'].value
            os.environ['Headlessmode'] = ws['B2'].value
            wb.close()
            configpath = os.environ['project_path'] + '\\Configurations\\GlobalConfiguration.ini'
            config=configparser.ConfigParser()
            config.read(configpath)
            for key in config['PATHS']:
                os.environ[key]=config['PATHS'][key]
           # cx_Oracle.init_oracle_client(lib_dir=r"C:instantclient")
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            folder_path = os.path.join(os.environ['project_path'], 'Reports', current_time)
            Path(folder_path).mkdir(parents=True, exist_ok=True)
            os.environ['report_path'] = folder_path
            logging.basicConfig(filename=os.path.join(folder_path, 'execution.log'), level=logging.INFO,
                                format='%(asctime)s - %(levelname)s - %(message)s')
        except Exception as e:
            print(f"Error in initialize_variable: {e}")
            sys.exit(1)


    def get_testcase_list(self):
        try:
            testcases_path = os.environ['project_path'] + '\\Configurations\\TestExecution.xlsx'
            df = pd.read_excel(testcases_path, sheet_name='TestCases')
            test_case_list = []
            for row in df.itertuples():
                if row.RunFlag == 'Y':
                    test_case=pd.read_excel(testcases_path,row.SheetName)
                    for row1 in test_case.itertuples():
                        if row1.RunFlag == 'Y':
                            testscriptpath=os.environ['Project_path']+row1.TestScriptPath
                            testscript=pd.read_excel(testscriptpath,'Sheet1')
                            for tc_row in testscript.itertuples():
                                testcasedetails=[]
                                if tc_row.testCaseName==row1.TestCaseName:
                                    testcasedetails.append(tc_row.TestCaseID)
                                    testcasedetails.append(tc_row.testCaseName)
                                    testcasedetails.append(tc_row.Function)
                                    testcasedetails.append(tc_row.Description)
                                    testcasedetails.append(tc_row.ExpectedResult)
                                    break
                    test_case_list.append(testcasedetails)
                    print(test_case_list)
            return test_case_list
        except Exception as e:
            logging.error(f"Error in get_testcase_list: {e}")
            return []
        
        if len(test_case_list)==0:
            logging.error("No test cases found with RunFlag 'Y'. Exiting.")
            sys.exit(1)
        else:
            del test_case
            del testset  
            del testscript
        return test_case_list